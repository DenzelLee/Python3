#!/usr/bin/env/python3
# -*- coding:utf-8 -*-
'''
Author:leo
Date&Time:2019-09-17 and 12:36
FileName:20190917_openpyxl库的excel读写.py
Description：...
'''
import datetime, time
import openpyxl

# 获取当前格式化-年月日：时分秒
nowTimee = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
# 获取当前格式化-年月日：时分秒
nowTime = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))

# 本地xlsx测试文件的绝对路径
filepath = r'用户信息数据库表.xlsx'
newfilepath = r'.\新用户信息数据库表.xlsx'

print("--------------------------------------------------------------")
print("一、创建和打印工作表名称")

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook(filepath)

# 从（workbook工作薄）中获取一个(sheet工作表)对象
sheets = wb.sheetnames
print(f"1.打印当前所有工作表：{sheets}\n2.当前工作表类型： {type(sheets)}")

# 创建一个sheet新工作表
mySheet = wb.create_sheet('newSheet')
print(f"3.创建新工作表成功，新工作表名称：{wb.sheetnames[-1]}")

# 读取指定的sheet工作表内容
count = 0
sheet1 = wb.get_sheet_by_name('用户信息数据库表')
sheet2 = wb['newSheet']
print(f"4.当前工作表清单：")
for sheet in wb:
    count +=1
    print(f"\t4-{count}.{sheet.title}")
print("--------------------------------------------------------------")

print("二、读取和操作单元格内容")

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook(filepath)

# 当前活跃的表单
ws = wb.active
print(f"1.当前活动的工作表（sheet）为：{ws}")

# 获取A列的第一个对象（第一行/第一列/第一个单元格）
print(f"2.寻找周杰伦对象和值：{ws['B9']}", ws['B9'].value)

# 通过对象获取值和行列坐标
c = ws['B18']
# 打印这个单元格对象所在的行列的数值和内容
print('3.第Row {}行, 第Column {}列 is 韩国演员：{}'.format(c.row, c.column, c.value))

# 获取单元格对象的所在列的行数和值
print('4.坐标值为{} is 韩国演员：{}\n'.format(c.coordinate, c.value))
print("--------------------------------------------------------------")

print("三、读取和操作单元格内容")
# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook(filepath)
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单

# 通过行列坐标，获取对象值（获取第一行第二列的单元格）
print(ws.cell(row=14, column=2),ws.cell(row=14, column=2).value)

# 循环打印第i行，第二列的全部值
for i in range(1, 5, 1):
    print(i, ws.cell(row=i, column=2).value)
print("--------------------------------------------------------------")
print("四、年龄大于50的用户，写入新表格第一列")
from openpyxl import Workbook
wb = openpyxl.load_workbook(filepath)
wb.create_sheet('newsheet')
ws = wb.get_sheet_by_name("用户信息数据库表")
count = 0
list_50 = []
for x in range(2, 20, 1):
    # for y in range(1,7,1):
        nid = ws.cell(row=x, column=1).value
        name = ws.cell(row=x, column=2).value
        phone = ws.cell(row=x, column=3).value
        sex = ws.cell(row=x, column=4).value
        age = ws.cell(row=x, column=5).value
        brith = str(ws.cell(row=x, column=6).value).split(" ")[0]
        info = f"{nid:>2},{name:<4},{phone:>11},{sex:>2},{age:>3},{brith}"
        # print(f"{nid:>2},{name:<4},{phone:>11},{sex:>2},{age:>3},{brith}")
        if age>50:
            print(f"{nid:>2},{name:<4},{phone:>11},{sex:>2},{age:>3},{brith}")
            list_50.append(info)

ws1 = wb.get_sheet_by_name("newsheet")
ws1 = wb.active
i = 1
for value in list_50:
    print(i, value)
    ws1.cell(i, 1, value)
    i += 1
wb.save(newfilepath)

print("-----------------------------------------------------------------")
print("五、单元格读写")
# 开始加载已有工作簿
wb = openpyxl.load_workbook(filepath)

# 遍历所有工作表名称
print(f"已存在工作表：{[sheetname for sheetname in wb.sheetnames]}")

# 实例化单个工作表
ws = wb.get_sheet_by_name("用户信息数据库表")
# 选中需要使用的工作表为活动工作表
ws = wb.active
print(f"当前活动工作表：{wb.active}")

# 打印单元格内容(返回元组类型）
print(f"打印指定行/列单元格内容：{ws.cell(row=10, column=2).value}")
print(f"打印指定第一格/第一行内容：  {ws['A1']}，{ws['1']}")
print(f"打印指定第一格/第一列内容：  {ws['A1']}，{ws['A']}")
print(f"打印指定多行/多列内容：  {ws['1':'20']},{ws['A1':'F20']}")
print(f"打印指定多行/多列内容：  {tuple(ws.rows)},{tuple(ws.columns)}")
print(f"打印统计最大行/最大列：  {ws.max_row},{ws.max_column}")

# sheet.rows为生成器, 表示每一行的数据，每一行又由一个tuple包裹，内嵌一个循环把包裹的每个单元格打印出来，就完成分行打印。
# sheet.columns类似，不过里面是每个tuple是每一列的单元格。
print([[cell.value for cell in row] for row in ws.rows])           # 按行打印每一条数据
print([[cell.value for cell in colum] for colum in ws.columns])    # 按列打印每一条数据

print("----实战练习：-------------------------------------------------------------")
print("-----------------------------------------------------------------")
print("六、年龄大于50的用户，写入新表格一对一")

# 实战演练：第一行第5个age字段>30，写入新的sheet工作表中
workbook = openpyxl.load_workbook(filepath)
worksheet = workbook.get_sheet_by_name("用户信息数据库表")
# 从（workbook工作薄）中获取一个(sheet工作表)对象
sheets = workbook.sheetnames
print(f"1.打印当前所有工作表：{sheets}\n2.当前工作表类型： {type(sheets)}")
line = [[cell.value for cell in row] for row in worksheet.rows]
for i in line:
    if int(line.index(i)) > 0 and int(i[4]) > 30:
        # 选中需要使用的工作表为活动工作表
        ws1 = workbook.active
        # ws1.title = "Sheet1"
        ws1 = workbook.get_sheet_by_name("Sheet1")  # ！！需要先active，再切换sheet，否则切换不了表格
        # print(f"当前活动工作表：{ws1}")
        ws1.append(i)  # append（[]），添加列表类型的数据

workbook.save(newfilepath)
wb.save(newfilepath)


