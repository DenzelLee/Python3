#!/usr/bin/env/python3
# -*- coding:utf-8 -*-
'''
Author:leo
Date&Time:2019-09-17 and 12:36
FileName:20190917_操作excel.py
Description：...
'''
from selenium import webdriver
from pprint import pprint
from time import sleep
import datetime, time

# 获取当前格式化-年月日：时分秒
nowTimee = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
# 获取当前格式化-年月日：时分秒
nowTime = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))

filepath = r'C:\Users\Administrator\Desktop\临时照片\测试数据《交易导入模板1》.xlsx'
import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook(filepath)

# 从工作薄中获取一个表单(sheet)对象
sheets = wb.sheetnames
print(f"当前已有工作表：{sheets}\n工作表类型： {type(sheets)}")

# 创建一个表单
mySheet = wb.create_sheet('mySheet')
print(f"新工作表：{wb.sheetnames}")

# 获取指定的表单
sheet1 = wb.get_sheet_by_name('Sheet1')
sheet2 = wb['mySheet']
count  = 0
for sheet in wb:
    count +=1
    print(f"{count}.{sheet.title}")
print("--------------------------------------------------------------")

import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook(filepath)
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单
print(ws)
print(ws['A1'])  # 获取A列的第一个对象
print(ws['A1'].value)

c = ws['B1']
# 打印这个单元格对象所在的行列的数值和内容
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
# 获取单元格对象的所在列的行数和值
print('Cell {} is {}\n'.format(c.coordinate, c.value))
print("--------------------------------------------------------------")

import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook(filepath)
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单

print(ws.cell(row=1, column=2))  # 获取第一行第二列的单元格
print(ws.cell(row=1, column=2).value)
for i in range(1, 8, 2):  # 获取1,3,4,7 行第二列的值
    print(i, ws.cell(row=i, column=2).value)